#!/usr/bin/env python3
"""
Exportador de Horário para Excel

Uso:
    python exportador_excel.py <horario.json> <saida.xlsx>

Gera um arquivo Excel com várias abas:
  - Resumo
  - Uma aba por turma (Turma_<id>)
  - Uma aba por professor (Prof_<nome>)
  - Uma aba por sala (Sala_<nome>)
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict


# Definição de turnos para planejamento
TEMPOS_MANHA = ["1", "2", "3", "4"]
TEMPOS_TARDE = ["5", "6", "7", "8", "9"]


def slots_planejamento(prof, config):
    """Retorna set de (dia, tempo) bloqueados para planejamento do professor."""
    plan = prof.get("planejamento")
    if not plan:
        return set()
    tempos_ids = [t["id"] for t in config["tempos"]]
    bloqueados = set()
    modalidade = plan.get("modalidade")
    if modalidade == "dia_inteiro":
        for dia in plan.get("dias", []):
            for tempo in tempos_ids:
                bloqueados.add((dia, tempo))
    elif modalidade == "dois_turnos":
        for ti in plan.get("turnos", []):
            tempos_turno = TEMPOS_MANHA if ti["turno"] == "manha" else TEMPOS_TARDE
            for tempo in tempos_turno:
                if tempo in tempos_ids:
                    bloqueados.add((ti["dia"], tempo))
    return bloqueados


# bordas padrão
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def cor_para_fill(cor_hex):
    """Recebe '#3B82F6' e retorna PatternFill compatível."""
    if not cor_hex or not cor_hex.startswith("#"):
        cor_hex = "#FFFFFF"
    # versão clarinha (com transparência simulada): mistura com branco
    h = cor_hex.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = int(r * 0.3 + 255 * 0.7)
    g = int(g * 0.3 + 255 * 0.7)
    b = int(b * 0.3 + 255 * 0.7)
    cor_clara = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=cor_clara, end_color=cor_clara, fill_type="solid")


def slot_to_dia_tempo(slot_key):
    """'seg_1' -> ('seg', '1')"""
    dia, tempo = slot_key.rsplit("_", 1)
    return dia, tempo


def montar_grid(horario, config, filtro_func, formatador_celula):
    """
    Cria uma matriz dia x tempo aplicando filtro_func(slot_info) e formatador_celula(slot_info).
    Retorna dict {(dia, tempo): (texto, area)}.
    """
    grid = {}
    for slot_key, aulas in horario.items():
        dia, tempo = slot_to_dia_tempo(slot_key)
        for turma, info in aulas.items():
            if filtro_func(turma, info):
                grid[(dia, tempo)] = (formatador_celula(turma, info), info.get("area", ""))
    return grid


def adicionar_aba_horario(wb, titulo, grid, config, areas, header_cell_text="", planejamento_slots=None):
    """Adiciona uma aba com a grade dia x tempo.
    
    planejamento_slots: set de (dia, tempo) que devem ser renderizados como PLANEJAMENTO."""
    if planejamento_slots is None:
        planejamento_slots = set()
    ws = wb.create_sheet(titulo[:31])  # max 31 chars

    dias = config["dias"]
    dias_label = {"seg": "Segunda", "ter": "Terça", "qua": "Quarta", "qui": "Quinta", "sex": "Sexta"}
    tempos = config["tempos"]

    # cabeçalho (linha 1)
    ws.cell(row=1, column=1, value=header_cell_text).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(dias) + 1)
    ws.cell(row=1, column=1).alignment = CENTER

    # linha 2: dias
    ws.cell(row=2, column=1, value="Tempo").font = Font(bold=True)
    ws.cell(row=2, column=1).fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    ws.cell(row=2, column=1).alignment = CENTER
    ws.cell(row=2, column=1).border = THIN_BORDER
    for i, dia in enumerate(dias):
        c = ws.cell(row=2, column=i + 2, value=dias_label.get(dia, dia))
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        c.alignment = CENTER
        c.border = THIN_BORDER

    # linhas: cada tempo
    cor_area = {nome: info.get("cor", "#CCCCCC") for nome, info in areas.items()}
    plan_fill = PatternFill(start_color="9CA3AF", end_color="9CA3AF", fill_type="solid")

    for i, tempo_info in enumerate(tempos):
        row = i + 3
        tempo_id = tempo_info["id"]
        rotulo = f"{tempo_id}º\n{tempo_info['inicio']}–{tempo_info['fim']}"
        c = ws.cell(row=row, column=1, value=rotulo)
        c.font = Font(bold=True, size=9)
        c.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        c.alignment = CENTER
        c.border = THIN_BORDER
        ws.row_dimensions[row].height = 55
        for j, dia in enumerate(dias):
            col = j + 2
            cell = ws.cell(row=row, column=col)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            # checagem prioritária: este slot é de planejamento?
            if (dia, tempo_id) in planejamento_slots:
                cell.value = "PLANEJAMENTO"
                cell.fill = plan_fill
                cell.font = Font(size=9, bold=True, color="FFFFFF")
                continue
            cell_data = grid.get((dia, tempo_id))
            if cell_data:
                texto, area = cell_data
                cell.value = texto
                if area in cor_area:
                    cell.fill = cor_para_fill(cor_area[area])

    # ajustar largura colunas
    ws.column_dimensions["A"].width = 12
    for i in range(len(dias)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 22


def aba_resumo(wb, resultado):
    entrada = resultado["entrada"]
    meta = resultado["metadados"]
    ws = wb.create_sheet("Resumo", 0)

    ws["A1"] = f"HORÁRIO ESCOLAR — {entrada['escola']['nome']}"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Ano letivo: {entrada['escola']['ano_letivo']}"
    ws["A2"].font = Font(size=11)

    row = 4
    ws.cell(row=row, column=1, value="ESTATÍSTICAS GERAIS").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Total de turmas:")
    ws.cell(row=row, column=2, value=len(entrada["turmas"]))
    row += 1
    ws.cell(row=row, column=1, value="Total de professores:")
    ws.cell(row=row, column=2, value=len(entrada["professores"]))
    row += 1
    n_salas = sum(len(a["salas"]) for a in entrada["areas"].values())
    ws.cell(row=row, column=1, value="Total de salas:")
    ws.cell(row=row, column=2, value=n_salas)
    row += 1
    ws.cell(row=row, column=1, value="Aulas alocadas:")
    ws.cell(row=row, column=2, value=meta["total_aulas_alocadas"])

    if meta.get("nao_alocadas"):
        row += 2
        ws.cell(row=row, column=1, value="⚠ AULAS NÃO ALOCADAS").font = Font(bold=True, color="EF4444")
        row += 1
        for na in meta["nao_alocadas"]:
            ws.cell(row=row, column=1, value=f"{na['turma']} - {na['disciplina']} ({na['professor']})")
            row += 1

    row += 2
    ws.cell(row=row, column=1, value="OCUPAÇÃO DE PROFESSORES").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Professor").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Aulas/semana").font = Font(bold=True)
    row += 1
    for prof, n in sorted(meta["ocupacao_professores"].items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=prof)
        ws.cell(row=row, column=2, value=n)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="OCUPAÇÃO DE SALAS").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Sala").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Aulas").font = Font(bold=True)
    ws.cell(row=row, column=3, value="% ocupação").font = Font(bold=True)
    row += 1
    total_slots = len(entrada["configuracao"]["dias"]) * entrada["configuracao"]["tempos_por_dia"]
    for sala, n in sorted(meta["ocupacao_salas"].items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=sala)
        ws.cell(row=row, column=2, value=n)
        ws.cell(row=row, column=3, value=f"{100*n/total_slots:.0f}%")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="PLANEJAMENTO DOS PROFESSORES").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Professor").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Modalidade").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Quando").font = Font(bold=True)
    row += 1
    dias_lbl = {"seg": "Seg", "ter": "Ter", "qua": "Qua", "qui": "Qui", "sex": "Sex"}
    turno_lbl = {"manha": "Manhã", "tarde": "Tarde"}
    for prof in entrada["professores"]:
        plan = prof.get("planejamento")
        if not plan:
            ws.cell(row=row, column=1, value=prof["nome"])
            ws.cell(row=row, column=2, value="(não definido)")
            ws.cell(row=row, column=3, value="—")
        else:
            ws.cell(row=row, column=1, value=prof["nome"])
            modalidade = plan.get("modalidade", "")
            if modalidade == "dia_inteiro":
                ws.cell(row=row, column=2, value="Dia inteiro")
                dias_str = ", ".join(dias_lbl.get(d, d) for d in plan.get("dias", []))
                ws.cell(row=row, column=3, value=dias_str)
            elif modalidade == "dois_turnos":
                ws.cell(row=row, column=2, value="Dois turnos")
                turnos_str = ", ".join(
                    f"{dias_lbl.get(t['dia'], t['dia'])} {turno_lbl.get(t['turno'], t['turno'])}"
                    for t in plan.get("turnos", [])
                )
                ws.cell(row=row, column=3, value=turnos_str)
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22


def exportar(resultado, caminho_saida):
    horario = resultado["horario"]
    entrada = resultado["entrada"]
    config = entrada["configuracao"]
    areas = entrada["areas"]

    wb = Workbook()
    # remover sheet default
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    aba_resumo(wb, resultado)

    # uma aba por turma
    for turma in entrada["turmas"]:
        tid = turma["id"]
        grid = montar_grid(
            horario, config,
            filtro_func=lambda t, info, _tid=tid: t == _tid,
            formatador_celula=lambda t, info: f"{info['disciplina']}\n{info['professor']}\n{info['sala']}"
        )
        adicionar_aba_horario(
            wb, f"Turma_{tid}", grid, config, areas,
            header_cell_text=f"Horário - Turma {tid}"
        )

    # uma aba por professor
    for prof in entrada["professores"]:
        pnome = prof["nome"]
        grid = montar_grid(
            horario, config,
            filtro_func=lambda t, info, _p=pnome: info["professor"] == _p,
            formatador_celula=lambda t, info: f"{t}\n{info['disciplina']}\n{info['sala']}"
        )
        nome_aba = "Prof_" + pnome.replace(" ", "")[:25]
        plan_slots = slots_planejamento(prof, config)
        adicionar_aba_horario(
            wb, nome_aba, grid, config, areas,
            header_cell_text=f"Horário - {pnome} ({prof['area']})",
            planejamento_slots=plan_slots,
        )

    # uma aba por sala
    todas_salas = []
    for area_nome, info in areas.items():
        for s in info["salas"]:
            todas_salas.append((s["nome"], area_nome, s["tipo"]))
    for sala_nome, area_nome, tipo in todas_salas:
        grid = montar_grid(
            horario, config,
            filtro_func=lambda t, info, _s=sala_nome: info["sala"] == _s,
            formatador_celula=lambda t, info: f"{t}\n{info['disciplina']}\n{info['professor']}"
        )
        nome_aba = "Sala_" + sala_nome.replace(" ", "")[:25]
        adicionar_aba_horario(
            wb, nome_aba, grid, config, areas,
            header_cell_text=f"Ocupação - {sala_nome} ({area_nome}, {tipo})"
        )

    wb.save(caminho_saida)
    print(f"Excel salvo em: {caminho_saida}")


def main():
    if len(sys.argv) < 3:
        print("Uso: python exportador_excel.py <horario.json> <saida.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], "r", encoding="utf-8") as f:
        resultado = json.load(f)
    exportar(resultado, sys.argv[2])


if __name__ == "__main__":
    main()
