import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

with open("/Users/kelsopalheta/Developer/Horario Escolar/output/grade_v2.json") as f:
    grade_raw = json.load(f)
with open("/Users/kelsopalheta/Developer/Horario Escolar/output/RESTRICOES.json") as f:
    R = json.load(f)

grade = {}
for k, v in grade_raw.items():
    dia, aula, turma = k.split("|")
    grade[(dia, int(aula[1:]), turma)] = (v["prof"], v["disc"])

dias = R["dias"]
turmas = sorted(set(k[2] for k in grade.keys()))

AREA_COLORS = {
    "Linguagens": "FFD9E1F2", "Matemática": "FFFCE4D6",
    "Natureza": "FFE2EFDA", "Humanas": "FFFFF2CC",
    "Itinerário": "FFEDEDED"
}
AREAS_MAP = {}
for area, discs in R["areas_disciplinas"].items():
    for d in discs: AREAS_MAP[d] = area

# Mapear sala
def get_sala(prof, disc, turma):
    # KELSO sempre Sala Arte (Arte + Eletiva)
    if prof == "KELSO": return "Sala Arte"
    if disc == "Técnico": return "Sala Info"
    if disc == "Ed. Física": return "Quadra"
    if disc == "Prát. Experimentais": return "Lab Multi"
    # MARÍLIA sempre Sala Info
    if prof == "MARÍLIA": return "Sala Info"

    profs_sala = {
        "ELIANA": "L1", "EUNICE": "L1",
        "JOANA": "L2", "REGINA": "L2",
        "VANESA": "L3",
        "LUCIDALVA": "M1", "MILTON": "M2", "LUIZ": "M3", "ROGÉRIO": "M3",
        "DIJANE": "N1", "EDMUNDO": "N2", "LOUISE": "N2", "RODRIGO": "N2",
        "ELENFLÁVIA": "H1", "LUCÍLIO": "H1",
        "ELINELSON": "H2", "CASIANO": "H2",
    }
    if prof == "JHONATAN":
        return "Quadra" if disc == "Ed. Física" else "L3"
    return profs_sala.get(prof, "?")

wb = Workbook()
wb.remove(wb.active)

thin = Side(border_style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="FF4472C4")
header_font = Font(bold=True, color="FFFFFFFF", size=11)

# Aba Resumo
ws = wb.create_sheet("Resumo")
ws['A1'] = "HORÁRIO ESCOLAR v2 — Pós-Otimização"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:F1')
ws.append([])
ws.append(["Métrica", "Valor"])
for c in ws[3]: c.font = header_font; c.fill = header_fill

# Contar
total = len(grade)

def prof_off(prof, dia, aula):
    for p in R["planejamento_pedagogico"].get(dia, []):
        if p["prof"] == prof:
            if p["periodo"] == "all": return True
            if p["periodo"] == "tarde" and aula >= 6: return True
    return False

viol = sum(1 for k, v in grade.items() if prof_off(v[0], k[0], k[1]))
confl_dict = defaultdict(list)
for k, v in grade.items():
    confl_dict[(k[0], k[1], v[0])].append(k[2])
confl = sum(1 for vv in confl_dict.values() if len(vv) > 1)

# Isoladas
def count_isoladas(g):
    cnt = 0
    for tt in turmas:
        for dia in dias:
            cur = None
            for aula in range(1, 10):
                if dia == "Quarta" and aula >= 8:
                    if cur and cur[2] == 1: cnt += 1
                    cur = None
                    continue
                e = g.get((dia, aula, tt))
                if e:
                    p, d = e
                    if cur and cur[0] == p and cur[1] == d:
                        cur = (p, d, cur[2]+1)
                    else:
                        if cur and cur[2] == 1: cnt += 1
                        cur = (p, d, 1)
                else:
                    if cur and cur[2] == 1: cnt += 1
                    cur = None
            if cur and cur[2] == 1: cnt += 1
    return cnt

iso = count_isoladas(grade)

ws.append(["Total slots", total])
ws.append(["Violações planejamento", viol])
ws.append(["Conflitos prof", confl])
ws.append(["Aulas isoladas", iso])
ws.append(["Total turmas", len(turmas)])

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 12

# Aba por turma com sala
for turma in turmas:
    safe = turma.replace("º", "").replace(" ", "_")[:30]
    ws = wb.create_sheet(f"T_{safe}")
    ws.column_dimensions['A'].width = 12
    for c in 'BCDEF': ws.column_dimensions[c].width = 26

    ws['A1'] = f"GRADE — {turma}"
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:F1')

    ws.append([])
    ws.append(["Aula/Dia"] + dias)
    for c in ws[3]: c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")

    for aula in range(1, 10):
        row_data = [f"Aula {aula}"]
        for dia in dias:
            if dia == "Quarta" and aula >= 8:
                row_data.append("🎯 CLUBES")
                continue
            e = grade.get((dia, aula, turma))
            if e:
                p, d = e
                sala = get_sala(p, d, turma)
                row_data.append(f"{p}\n{d}\n📍 {sala}")
            else:
                row_data.append("—")
        ws.append(row_data)
        cur_row = ws.max_row
        for i, dia in enumerate(dias, 2):
            e = grade.get((dia, aula, turma)) if not (dia == "Quarta" and aula >= 8) else None
            if e:
                area = AREAS_MAP.get(e[1], "Itinerário")
                ws.cell(cur_row, i).fill = PatternFill("solid", fgColor=AREA_COLORS.get(area))
            ws.cell(cur_row, i).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            ws.cell(cur_row, i).border = border
            ws.cell(cur_row, i).font = Font(size=9)
        ws.row_dimensions[cur_row].height = 50

# Aba por professor
ws_p = wb.create_sheet("Por Professor")
ws_p.column_dimensions['A'].width = 14
ws_p.column_dimensions['B'].width = 12
for c in 'CDEFG': ws_p.column_dimensions[c].width = 26

profs = sorted(set(v[0] for v in grade.values()))
for prof in profs:
    ws_p.append([])
    ws_p.append([f"PROFESSOR: {prof}"])
    ws_p.cell(ws_p.max_row, 1).font = Font(bold=True, size=12)
    ws_p.append(["Aula/Dia"] + dias)
    for c in ws_p[ws_p.max_row]: c.font = header_font; c.fill = header_fill

    for aula in range(1, 10):
        row_data = [f"Aula {aula}"]
        for dia in dias:
            if dia == "Quarta" and aula >= 8:
                row_data.append("CLUBES")
                continue
            found = None
            for (d, a, t), (p, dc) in grade.items():
                if d == dia and a == aula and p == prof:
                    sala = get_sala(p, dc, t)
                    found = f"{t}\n{dc}\n📍 {sala}"
                    break
            if not found and prof_off(prof, dia, aula):
                found = "📚 PLANEJAMENTO"
            row_data.append(found or "—")
        ws_p.append(row_data)
        cr = ws_p.max_row
        for i in range(2, 7):
            ws_p.cell(cr, i).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            ws_p.cell(cr, i).border = border
            ws_p.cell(cr, i).font = Font(size=8)
            if ws_p.cell(cr, i).value == "📚 PLANEJAMENTO":
                ws_p.cell(cr, i).fill = PatternFill("solid", fgColor="FFFFE699")
        ws_p.row_dimensions[cr].height = 45

# Aba salas
ws_s = wb.create_sheet("Por Sala")
salas = ["Sala L1", "Sala L2", "Sala L3", "Sala Leitura",
         "Sala M1", "Sala M2", "Sala M3",
         "Sala N1", "Sala N2", "Lab Multi",
         "Sala H1", "Sala H2", "Auditório",
         "Sala Arte", "Sala Info", "Quadra"]

ws_s.column_dimensions['A'].width = 16
for c in 'BCDEF': ws_s.column_dimensions[c].width = 24

for sala in salas:
    ws_s.append([])
    ws_s.append([f"SALA: {sala}"])
    ws_s.cell(ws_s.max_row, 1).font = Font(bold=True, size=12)
    ws_s.append(["Aula/Dia"] + dias)
    for c in ws_s[ws_s.max_row]: c.font = header_font; c.fill = header_fill

    for aula in range(1, 10):
        row = [f"Aula {aula}"]
        for dia in dias:
            if dia == "Quarta" and aula >= 8:
                row.append("CLUBES")
                continue
            found = None
            for (d, a, t), (p, dc) in grade.items():
                if d == dia and a == aula:
                    sala_calc = get_sala(p, dc, t)
                    # Mapeia abreviações p/ comparar
                    sala_full = {"L1":"Sala L1","L2":"Sala L2","L3":"Sala L3","M1":"Sala M1","M2":"Sala M2","M3":"Sala M3","N1":"Sala N1","N2":"Sala N2","H1":"Sala H1","H2":"Sala H2"}.get(sala_calc, sala_calc)
                    if sala_full == sala:
                        found = f"{t}\n{p}/{dc}"
                        break
            row.append(found or "—")
        ws_s.append(row)
        cr = ws_s.max_row
        for i in range(2, 7):
            ws_s.cell(cr, i).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            ws_s.cell(cr, i).border = border
            ws_s.cell(cr, i).font = Font(size=8)
        ws_s.row_dimensions[cr].height = 40

wb.save("/Users/kelsopalheta/Developer/Horario Escolar/output/Horario_Escolar_v2.xlsx")
print("✓ Horario_Escolar_v2.xlsx criado")
print(f"  Violações: {viol} | Conflitos: {confl} | Isoladas: {iso}")
