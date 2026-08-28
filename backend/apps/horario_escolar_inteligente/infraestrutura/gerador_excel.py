"""
Gerador de Relatórios Excel Multi-Aba (src/infraestrutura/gerador_excel.py)
Geração profissional com openpyxl formatado e estilizado por área BNCC
"""
import io
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.horario_escolar_inteligente.dominio.entidades import GradeEscolar, Slot


class GeradorExcelRelatorio:
    """
    Gera pasta de trabalho Excel (.xlsx) com múltiplas abas:
    1. Quadro por Turma
    2. Quadro por Professor
    3. Ocupação de Salas Temáticas
    """

    CORES_AREA = {
        "Linguagens": "D9E1F2",   # Azul pastel
        "Matemática": "FCE4D6",   # Laranja pastel
        "Natureza": "E2EFDA",     # Verde pastel
        "Humanas": "FFF2CC",      # Amarelo pastel
        "Especiais": "EDEDED",    # Cinza claro
    }

    DIAS = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    AULAS = list(range(1, 10))

    def gerar_bytes(self, slots: List[Slot]) -> bytes:
        """Gera o arquivo Excel em memória e retorna os bytes."""
        wb = openpyxl.Workbook()
        # Remover aba padrão
        wb.remove(wb.active)

        self._criar_aba_turmas(wb, slots)
        self._criar_aba_professores(wb, slots)
        self._criar_aba_salas(wb, slots)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    def _criar_aba_turmas(self, wb: openpyxl.Workbook, slots: List[Slot]):
        ws = wb.create_sheet(title="Quadro por Turma")
        ws.views.sheetView[0].showGridLines = True

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # Cabeçalhos
        headers = ["Turma", "Dia", "Aula", "Disciplina", "Professor", "Sala Temática"]
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Linhas
        row_num = 2
        for s in slots:
            ws.cell(row=row_num, column=1, value=s.turma).border = thin_border
            ws.cell(row=row_num, column=2, value=s.dia).border = thin_border
            ws.cell(row=row_num, column=3, value=s.aula).border = thin_border
            ws.cell(row=row_num, column=4, value=s.disc).border = thin_border
            ws.cell(row=row_num, column=5, value=s.prof).border = thin_border
            ws.cell(row=row_num, column=6, value=s.sala).border = thin_border
            row_num += 1

        self._ajustar_largura_colunas(ws)

    def _criar_aba_professores(self, wb: openpyxl.Workbook, slots: List[Slot]):
        ws = wb.create_sheet(title="Quadro por Professor")
        ws.views.sheetView[0].showGridLines = True

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = ["Professor", "Dia", "Aula", "Turma", "Disciplina", "Sala"]
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        profs_ordenados = sorted(slots, key=lambda x: (x.prof, x.dia, x.aula))
        for row_num, s in enumerate(profs_ordenados, 2):
            ws.cell(row=row_num, column=1, value=s.prof)
            ws.cell(row=row_num, column=2, value=s.dia)
            ws.cell(row=row_num, column=3, value=s.aula)
            ws.cell(row=row_num, column=4, value=s.turma)
            ws.cell(row=row_num, column=5, value=s.disc)
            ws.cell(row=row_num, column=6, value=s.sala)

        self._ajustar_largura_colunas(ws)

    def _criar_aba_salas(self, wb: openpyxl.Workbook, slots: List[Slot]):
        ws = wb.create_sheet(title="Salas Temáticas")
        ws.views.sheetView[0].showGridLines = True

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = ["Sala Temática", "Dia", "Aula", "Turma", "Professor", "Disciplina"]
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        salas_ordenadas = sorted(slots, key=lambda x: (x.sala, x.dia, x.aula))
        for row_num, s in enumerate(salas_ordenadas, 2):
            ws.cell(row=row_num, column=1, value=s.sala)
            ws.cell(row=row_num, column=2, value=s.dia)
            ws.cell(row=row_num, column=3, value=s.aula)
            ws.cell(row=row_num, column=4, value=s.turma)
            ws.cell(row=row_num, column=5, value=s.prof)
            ws.cell(row=row_num, column=6, value=s.disc)

        self._ajustar_largura_colunas(ws)

    def _ajustar_largura_colunas(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
